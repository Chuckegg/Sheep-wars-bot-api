#!/usr/bin/env python3
"""
Convert stats.xlsx to stats.db SQLite database.
This script extracts all data from the Excel file and migrates it to SQLite.
"""

import sqlite3
import openpyxl
from pathlib import Path
import sys

EXCEL_FILE = Path(__file__).parent / "stats.xlsx"
DB_FILE = Path(__file__).parent / "stats.db"

def create_database_schema(conn):
    """Create the database schema for stats storage."""
    cursor = conn.cursor()
    
    # User stats table - stores all stat values for each user
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS user_stats (
            username TEXT NOT NULL,
            stat_name TEXT NOT NULL,
            lifetime REAL DEFAULT 0,
            session REAL DEFAULT 0,
            daily REAL DEFAULT 0,
            yesterday REAL DEFAULT 0,
            monthly REAL DEFAULT 0,
            PRIMARY KEY (username, stat_name)
        )
    ''')
    
    # User metadata table - stores level, icon, colors, etc.
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS user_meta (
            username TEXT PRIMARY KEY,
            level INTEGER DEFAULT 0,
            icon TEXT DEFAULT '',
            ign_color TEXT DEFAULT NULL,
            guild_tag TEXT DEFAULT NULL,
            guild_hex TEXT DEFAULT NULL
        )
    ''')
    
    # Create indexes for faster lookups
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_username ON user_stats(username)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_stat_name ON user_stats(stat_name)')
    
    conn.commit()
    print("[DB] Database schema created successfully")

def extract_excel_data(excel_path):
    """Extract all data from stats.xlsx."""
    if not excel_path.exists():
        print(f"[ERROR] Excel file not found: {excel_path}")
        sys.exit(1)
    
    print(f"[EXCEL] Loading {excel_path}...")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    all_users_data = {}
    
    for sheet_name in wb.sheetnames:
        if sheet_name == "Template":
            continue
        
        sheet = wb[sheet_name]
        username = sheet_name
        
        # Initialize user data
        user_data = {
            'stats': {},
            'meta': {}
        }
        
        # Extract metadata from first few rows
        # Row 1: Username (A1), Level value (B1)
        # Row 2: Level label (A2), Level value with icon (B2) 
        # Row 3: Icon (A3), Icon value (B3)
        # Row 4: IGN Color (A4), Color value (B4)
        # Row 5: Guild Tag (A5), Tag value (B5)
        # Row 6: Guild Hex (A6), Hex value (B6)
        
        try:
            level_val = sheet.cell(row=2, column=2).value
            if level_val and isinstance(level_val, (int, float)):
                user_data['meta']['level'] = int(level_val)
            else:
                user_data['meta']['level'] = 0
                
            icon_val = sheet.cell(row=3, column=2).value
            user_data['meta']['icon'] = str(icon_val) if icon_val else ''
            
            color_val = sheet.cell(row=4, column=2).value
            user_data['meta']['ign_color'] = str(color_val) if color_val else None
            
            guild_tag = sheet.cell(row=5, column=2).value
            user_data['meta']['guild_tag'] = str(guild_tag) if guild_tag else None
            
            guild_hex = sheet.cell(row=6, column=2).value
            user_data['meta']['guild_hex'] = str(guild_hex) if guild_hex else None
        except Exception as e:
            print(f"[WARN] Error extracting metadata for {username}: {e}")
        
        # Extract stats starting from row 2 (row 1 has headers)
        # Column A: Stat name
        # Column B: Lifetime value
        # Column C: Session Delta
        # Column D: Session value (snapshot)
        # Column E: Daily Delta
        # Column F: Daily value (snapshot)
        # Column G: Yesterday Delta
        # Column H: Yesterday value (snapshot)
        # Column I: Monthly Delta
        # Column J: Monthly value (snapshot)
        
        for row_idx in range(2, sheet.max_row + 1):
            stat_name_cell = sheet.cell(row=row_idx, column=1).value
            
            if not stat_name_cell:
                continue
            
            stat_name = str(stat_name_cell).strip()
            if not stat_name:
                continue
            
            # Extract values for each period
            lifetime_val = sheet.cell(row=row_idx, column=2).value  # Column B
            session_val = sheet.cell(row=row_idx, column=4).value   # Column D
            daily_val = sheet.cell(row=row_idx, column=6).value     # Column F
            yesterday_val = sheet.cell(row=row_idx, column=8).value # Column H
            monthly_val = sheet.cell(row=row_idx, column=10).value  # Column J
            
            # Convert to float, default to 0 if None or invalid
            def safe_float(val):
                try:
                    return float(val) if val is not None else 0.0
                except (ValueError, TypeError):
                    return 0.0
            
            user_data['stats'][stat_name] = {
                'lifetime': safe_float(lifetime_val),
                'session': safe_float(session_val),
                'daily': safe_float(daily_val),
                'yesterday': safe_float(yesterday_val),
                'monthly': safe_float(monthly_val)
            }
        
        all_users_data[username] = user_data
        print(f"[EXCEL] Extracted data for {username}: {len(user_data['stats'])} stats")
    
    wb.close()
    print(f"[EXCEL] Extraction complete. Total users: {len(all_users_data)}")
    return all_users_data

def insert_data_to_db(conn, all_users_data):
    """Insert all extracted data into the database."""
    cursor = conn.cursor()
    
    stats_inserted = 0
    users_inserted = 0
    
    for username, user_data in all_users_data.items():
        # Insert metadata
        meta = user_data['meta']
        cursor.execute('''
            INSERT OR REPLACE INTO user_meta (username, level, icon, ign_color, guild_tag, guild_hex)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (
            username,
            meta.get('level', 0),
            meta.get('icon', ''),
            meta.get('ign_color'),
            meta.get('guild_tag'),
            meta.get('guild_hex')
        ))
        users_inserted += 1
        
        # Insert stats
        for stat_name, periods in user_data['stats'].items():
            # Make stat_name lowercase for consistency
            stat_name_lower = stat_name.lower()
            cursor.execute('''
                INSERT OR REPLACE INTO user_stats (username, stat_name, lifetime, session, daily, yesterday, monthly)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (
                username,
                stat_name_lower,
                periods['lifetime'],
                periods['session'],
                periods['daily'],
                periods['yesterday'],
                periods['monthly']
            ))
            stats_inserted += 1
    
    conn.commit()
    print(f"[DB] Inserted {users_inserted} users and {stats_inserted} stat records")

def verify_conversion(conn, all_users_data):
    """Verify that all data was converted correctly."""
    cursor = conn.cursor()
    
    print("\n[VERIFY] Checking data integrity...")
    
    # Check user count
    cursor.execute('SELECT COUNT(DISTINCT username) FROM user_stats')
    db_user_count = cursor.fetchone()[0]
    excel_user_count = len(all_users_data)
    
    print(f"  Users: Excel={excel_user_count}, DB={db_user_count} {'✓' if db_user_count == excel_user_count else '✗'}")
    
    # Sample check: verify a few random stats
    errors = 0
    for username in list(all_users_data.keys())[:3]:  # Check first 3 users
        for stat_name in list(all_users_data[username]['stats'].keys())[:3]:  # Check first 3 stats
            cursor.execute('''
                SELECT lifetime, session, daily, yesterday, monthly 
                FROM user_stats 
                WHERE username = ? AND stat_name = ?
            ''', (username, stat_name))
            
            result = cursor.fetchone()
            if result:
                excel_data = all_users_data[username]['stats'][stat_name]
                db_data = {'lifetime': result[0], 'session': result[1], 'daily': result[2], 
                          'yesterday': result[3], 'monthly': result[4]}
                
                if excel_data != db_data:
                    print(f"  [WARN] Mismatch for {username}.{stat_name}")
                    errors += 1
    
    if errors == 0:
        print("  Sample verification: All checks passed ✓")
    else:
        print(f"  Sample verification: {errors} errors found ✗")
    
    print("\n[VERIFY] Conversion complete!")

def main():
    """Main conversion process."""
    print("=" * 60)
    print("Excel to SQLite Conversion Tool")
    print("=" * 60)
    
    # Check if database already exists
    if DB_FILE.exists():
        response = input(f"\n[WARN] {DB_FILE} already exists. Overwrite? (yes/no): ")
        if response.lower() != 'yes':
            print("[ABORT] Conversion cancelled.")
            sys.exit(0)
        DB_FILE.unlink()
        print(f"[DB] Deleted existing {DB_FILE}")
    
    # Create database and schema
    conn = sqlite3.connect(DB_FILE)
    create_database_schema(conn)
    
    # Extract data from Excel
    all_users_data = extract_excel_data(EXCEL_FILE)
    
    # Insert data into database
    insert_data_to_db(conn, all_users_data)
    
    # Verify conversion
    verify_conversion(conn, all_users_data)
    
    # Close connection
    conn.close()
    
    print("\n" + "=" * 60)
    print(f"SUCCESS: Data converted to {DB_FILE}")
    print("=" * 60)
    print("\nNext steps:")
    print("  1. Keep stats.xlsx as a backup")
    print("  2. Update discord_bot.py and api_get.py to use SQLite")
    print("  3. Test the bot with the new database")

if __name__ == "__main__":
    main()
